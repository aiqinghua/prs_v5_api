# _*_ coding : utf-8 _*_
# @Time : 2024/4/14 16:05
# @Author : aiqinghua
# @File : excle_utli
# @Project : prs_v5
from openpyxl import load_workbook


class HandleExcle:
    def __init__(self, excle_file_name, excle_sheet_name):
        """
        :param excle_file_name: excle文件名
        :param excle_sheet_name: sheet页名称
        """
        self.excle_file_name = excle_file_name
        self.excle_sheet_name = excle_sheet_name

    def open(self):
        """
        获取工作簿
        :return:
        """
        self.wb = load_workbook(self.excle_file_name)
        self.sheet_name = self.wb[self.excle_sheet_name]
        # self.sheet_name = self.wb.worksheets[0]

    def close(self):
        self.wb.close()

    def read_data(self):
        """
        读取excle方法
        :return:
        """
        self.open()
        rows = list(self.sheet_name.rows)
        titles = []
        for t in rows[0]:
            # 遍历获取表头数据
            title = t.value
            titles.append(title)
        # 获取表头也可以简写为：titles = [i.value for i in rows[0]]
        cases = []
        for row in rows[1:]:
            # 遍历除第一行的其他每行数据
            case = []
            for r in row:
                # 遍历每列数据
                case.append(r.value)
            # 获取is_execute列的索引
            is_execute_index = titles.index("is_execute")
            # 如果is_execute列的值为0，则跳过该行不执行
            if case[is_execute_index] == 0:
                continue
            # 将表头数据与每列的值合并成字典
            cases.append(dict(zip(titles, case)))
        self.close()
        return cases

    def write_data(self, row, column, value):
        """
        写入excle数据
        :param row:写入的行
        :param column:写入的列
        :param value:写入的值
        :return:
        """
        self.open()
        # 写入数据
        self.sheet_name.cell(row=row, column=column, value=value)
        self.wb.save(self.excle_file_name)
        self.close()


if __name__ == '__main__':
#     # print(os.path.dirname(os.path.realpath(__file__)))
    test = HandleExcle(r"C:\data\python\prs_v5\data\apicase.xlsx", "allrisk")
    # print(test.read_data())
    print(test.read_data())
    # for i in test.read_data():
    #     print(i)
    # test.write_date(row=1, column=8, value="res")